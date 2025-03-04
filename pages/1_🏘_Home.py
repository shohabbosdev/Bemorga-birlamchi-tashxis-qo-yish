import sqlite3
import streamlit as st
import pandas as pd
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from contextlib import contextmanager
from streamlit_tags import st_tags

# Thread-safe database connection management
@contextmanager
def get_connection():
    conn = None
    try:
        conn = sqlite3.connect('data/diagnosis_data.db', timeout=10, check_same_thread=False)
        yield conn
    finally:
        if conn:
            conn.close()

# Jadvalni yaratish
def create_tables():
    try:
        with get_connection() as conn:
            cursor = conn.cursor()
            
            # 'diseases' jadvalini yaratish
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS diseases (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    name TEXT NOT NULL
                )
            """)

            # 'symptom_groups' jadvalini yaratish
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS symptom_groups (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    disease_id INTEGER NOT NULL,
                    group_name TEXT NOT NULL,
                    FOREIGN KEY(disease_id) REFERENCES diseases(id)
                )
            """)

            # 'symptoms' jadvalini yaratish
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS symptoms (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    group_id INTEGER NOT NULL,
                    symptom_name TEXT NOT NULL,
                    FOREIGN KEY(group_id) REFERENCES symptom_groups(id)
                )
            """)

            # 'disease_symptoms' jadvalini yaratish
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS disease_symptoms (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    disease_id INTEGER NOT NULL,
                    symptom_id INTEGER NOT NULL,
                    value INTEGER NOT NULL,
                    FOREIGN KEY(disease_id) REFERENCES diseases(id),
                    FOREIGN KEY(symptom_id) REFERENCES symptoms(id)
                )
            """)

            conn.commit()
    except sqlite3.Error as e:
        st.error(f"Jadvallarni yaratishda xatolik yuz berdi: {str(e)}")

# Jadvalni yaratish funksiyasini dastur boshlang'ichiga qo'shish
create_tables()

# Ma'lumotlarni olish funksiyasi
def get_all_data():
    with get_connection() as conn:
        cursor = conn.cursor()
        
        # Kasalliklarni olish
        diseases = cursor.execute("SELECT id, name FROM diseases").fetchall()
        
        # Simptom guruhlarini olish
        groups = cursor.execute(""" 
            SELECT sg.id, sg.disease_id, sg.group_name, d.name as disease_name 
            FROM symptom_groups sg 
            JOIN diseases d ON sg.disease_id = d.id
        """).fetchall()
        
        # Simptomlarni olish
        symptoms = cursor.execute(""" 
            SELECT s.id, s.group_id, s.symptom_name, sg.group_name 
            FROM symptoms s 
            JOIN symptom_groups sg ON s.group_id = sg.id
        """).fetchall()
        
        # Qiymatlarni olish va to'g'ri filtrlash
        values = cursor.execute(""" 
            SELECT ds.id, ds.disease_id, ds.symptom_id, ds.value,
                d.name as disease_name, s.symptom_name, sg.group_name
            FROM disease_symptoms ds 
            JOIN diseases d ON ds.disease_id = d.id 
            JOIN symptoms s ON ds.symptom_id = s.id 
            JOIN symptom_groups sg ON s.group_id = sg.id
        """).fetchall()
        
        return diseases, groups, symptoms, values
    
# Keshni tozalash funksiyasi
def clear_cache():
    """Ma'lumotlar bazasidan keyin keshni tozalash"""
    if 'cached_data' in st.session_state:
        del st.session_state['cached_data']
    refresh_data()

def update_data(table, field, value, condition_field, condition_value):
    try:
        with get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute(f"UPDATE {table} SET {field}=? WHERE {condition_field}=?", 
                        (value, condition_value))
            conn.commit()
        clear_cache()  # Keshni tozalash
        return True
    except sqlite3.Error as e:
        st.error(f"Xatolik yuz berdi: {str(e)}")
        return False

def delete_data(table, condition_field, condition_value):
    try:
        with get_connection() as conn:
            cursor = conn.cursor()
            # Bog'liq ma'lumotlarni o'chirish
            if table == "diseases":
                # Kasallik o'chirilganda unga bog'liq barcha ma'lumotlarni o'chirish
                cursor.execute("SELECT id FROM symptom_groups WHERE disease_id=?", (condition_value,))
                group_ids = [row[0] for row in cursor.fetchall()]
                
                for group_id in group_ids:
                    cursor.execute("DELETE FROM symptoms WHERE group_id=?", (group_id,))
                
                cursor.execute("DELETE FROM symptom_groups WHERE disease_id=?", (condition_value,))
                cursor.execute("DELETE FROM disease_symptoms WHERE disease_id=?", (condition_value,))
            
            elif table == "symptom_groups":
                # Guruh o'chirilganda unga bog'liq simptomlarni o'chirish
                cursor.execute("SELECT id FROM symptoms WHERE group_id=?", (condition_value,))
                symptom_ids = [row[0] for row in cursor.fetchall()]
                
                for symptom_id in symptom_ids:
                    cursor.execute("DELETE FROM disease_symptoms WHERE symptom_id=?", (symptom_id,))
                
                cursor.execute("DELETE FROM symptoms WHERE group_id=?", (condition_value,))
            
            # Asosiy jadvaldan o'chirish
            cursor.execute(f"DELETE FROM {table} WHERE {condition_field}=?", (condition_value,))
            conn.commit()
            
        clear_cache()  # Keshni tozalash
        return True
    except sqlite3.Error as e:
        st.error(f"Xatolik yuz berdi: {str(e)}")
        return False

def update_symptom_value(disease_id, symptom_id, new_value):
    """Simptom qiymatini yangilash"""
    try:
        with get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute(""" 
                INSERT OR REPLACE INTO disease_symptoms (disease_id, symptom_id, value) 
                VALUES (?, ?, ?)
            """, (disease_id, symptom_id, new_value))
            conn.commit()
        clear_cache()
        return True
    except sqlite3.Error as e:
        st.error(f"Qiymatni yangilashda xatolik: {str(e)}")
        return False

# Session state orqali ma'lumotlarni saqlash
if 'cached_data' not in st.session_state:
    st.session_state.cached_data = get_all_data()

def refresh_data():
    st.session_state.cached_data = get_all_data()

# Kasallik va guruhni tekshirish
def check_and_insert_disease(conn, disease_name):
    cursor = conn.cursor()
    cursor.execute("SELECT id FROM diseases WHERE name=?", (disease_name,))
    disease_record = cursor.fetchone()
    if not disease_record:
        cursor.execute("INSERT INTO diseases (name) VALUES (?)", (disease_name,))
        conn.commit()
        return cursor.lastrowid
    return disease_record[0]

# Guruhni tekshirish va qo'shish
def check_and_insert_group(conn, disease_id, group_name):
    cursor = conn.cursor()
    cursor.execute("SELECT id FROM symptom_groups WHERE disease_id=? AND group_name=?", (disease_id, group_name))
    group_record = cursor.fetchone()
    if not group_record:
        cursor.execute("INSERT INTO symptom_groups (disease_id, group_name) VALUES (?, ?)", (disease_id, group_name))
        conn.commit()
        return cursor.lastrowid
    return group_record[0]

# Simptomni tekshirish va qo'shish
def check_and_insert_symptom(conn, group_id, symptom_name):
    cursor = conn.cursor()
    cursor.execute("SELECT id FROM symptoms WHERE group_id=? AND symptom_name=?", (group_id, symptom_name))
    symptom_record = cursor.fetchone()
    if not symptom_record:
        cursor.execute("INSERT INTO symptoms (group_id, symptom_name) VALUES (?, ?)", (group_id, symptom_name))
        conn.commit()
        return cursor.lastrowid
    return symptom_record[0]

# Qiymatlarni saqlash
def save_symptom_value(conn, disease_id, symptom_id, value):
    cursor = conn.cursor()
    cursor.execute("SELECT id FROM disease_symptoms WHERE disease_id=? AND symptom_id=?", (disease_id, symptom_id))
    existing_value = cursor.fetchone()
    if existing_value:
        cursor.execute("UPDATE disease_symptoms SET value=? WHERE disease_id=? AND symptom_id=?", (value, disease_id, symptom_id))
    else:
        cursor.execute("INSERT INTO disease_symptoms (disease_id, symptom_id, value) VALUES (?, ?, ?)", (disease_id, symptom_id, value))
    conn.commit()

# Tahrirlash oynasi
def edit_tab():
    st.markdown("## Ma'lumotlarni tahrirlash")

    edit_type = st.selectbox("Nimani tahrirlash kerak?", 
                            ["Kasalliklar", "Simptom guruhlari", "Simptomlar", "Qiymatlar"])

    diseases, groups, symptoms, values = st.session_state.cached_data
    
    if edit_type == "Kasalliklar":
        st.markdown("### Kasalliklarni qo'shish, o'zgartirish yoki o'chirish")
        
        # Kasalliklar ro'yxatini olish
        disease_dict = {d[1]: d[0] for d in diseases}
        
        # Kasallikni tanlash
        selected_disease = st.selectbox("Tahrir qilmoqchi bo'lgan kasallikni tanlang", list(disease_dict.keys()))
        
        # Kasallikni yangilash
        if selected_disease:
            disease_id = disease_dict[selected_disease]
            new_name = st.text_input("Yangi kasallik nomini kiriting", value=selected_disease)
            if st.button("Yangilash", key="update_disease", icon='🔄', use_container_width=True):
                if update_data("diseases", "name", new_name, "id", disease_id):
                    st.success("✅ Kasallik nomi yangilandi!")

            # Kasallikni o'chirish
            if st.button("O'chirish", key="delete_disease", type='primary', icon='❌', use_container_width=True):
                if delete_data("diseases", "id", disease_id):
                    st.success("✅ Kasallik o'chirildi!")
        
        # Kasallikni qo'shish
        new_disease_name = st.text_input("Yangi kasallik nomini kiriting", key="new_disease_name")
        if st.button("Kasallik qo'shish"):
            if new_disease_name:
                with get_connection() as conn:
                    disease_id = check_and_insert_disease(conn, new_disease_name)
                st.success(f"✅ {new_disease_name} kasalligi qo'shildi!")

    elif edit_type == "Simptom guruhlari":
        st.markdown("### Simptom guruhlarini qo'shish, o'zgartirish yoki o'chirish")
        
        # Simptom guruhlarini olish
        group_dict = {group[2]: group[0] for group in groups}
        
        # Simptom guruhini tanlash
        selected_group = st.selectbox("Tahrir qilmoqchi bo'lgan simptom guruhini tanlang", list(group_dict.keys()))
        
        # Simptom guruhini yangilash
        if selected_group:
            group_id = group_dict[selected_group]
            new_name = st.text_input("Yangi guruh nomini kiriting", value=selected_group)
            if st.button("Yangilash", key="update_group",icon='🔄', use_container_width=True):
                if update_data("symptom_groups", "group_name", new_name, "id", group_id):
                    st.success("✅ Simptom guruhi nomi yangilandi!")

            # Simptom guruhini o'chirish
            if st.button("O'chirish", key="delete_group", type='primary', icon='❌', use_container_width=True):
                if delete_data("symptom_groups", "id", group_id):
                    st.success("✅ Simptom guruhi o'chirildi!")
        
        # Simptom guruhini qo'shish
        new_group_name = st.text_input("Yangi simptom guruhini kiriting", key="new_group_name")
        if st.button("Simptom guruhi qo'shish"):
            if new_group_name:
                with get_connection() as conn:
                    disease_id = st.selectbox("Kasallikni tanlang", [d[1] for d in diseases])
                    group_id = check_and_insert_group(conn, disease_id, new_group_name)
                st.success(f"✅ {new_group_name} guruh qo'shildi!")

    elif edit_type == "Simptomlar":
        st.markdown("### Simptomlarni qo'shish, o'zgartirish yoki o'chirish")
        
        symptom_dict = {symptom[2]: symptom[0] for symptom in symptoms}
        
        # Simptomni tanlash
        selected_symptom = st.selectbox("Tahrir qilmoqchi bo'lgan simptomni tanlang", list(symptom_dict.keys()))
        
        # Simptomni yangilash
        if selected_symptom:
            symptom_id = symptom_dict[selected_symptom]
            new_name = st.text_input("Yangi simptom nomini kiriting", value=selected_symptom)
            if st.button("Yangilash", key="update_symptom", icon='🔄', use_container_width=True):
                if update_data("symptoms", "symptom_name", new_name, "id", symptom_id):
                    st.success("✅ Simptom nomi yangilandi!")

            # Simptomni o'chirish
            if st.button("O'chirish", key="delete_symptom", type='primary', icon='❌', use_container_width=True):
                if delete_data("symptoms", "id", symptom_id):
                    st.success("✅ Simptom o'chirildi!")
        
        # Simptomni qo'shish
        new_symptom_name = st.text_input("Yangi simptomni kiriting", key="new_symptom_name")
        if st.button("Simptom qo'shish"):
            if new_symptom_name:
                with get_connection() as conn:
                    group_id = st.selectbox("Simptom guruhini tanlang", [g[2] for g in groups])
                    symptom_id = check_and_insert_symptom(conn, group_id, new_symptom_name)
                st.success(f"✅ {new_symptom_name} simptom qo'shildi!")

    elif edit_type == "Qiymatlar":
        st.markdown("### Qiymatlarni qo'shish, o'zgartirish yoki o'chirish")
        
        values_dict = {f"{v[4]} - {v[5]} ({v[6]})": v[0] for v in values}
        
        # Qiymatni tanlash
        selected_value = st.selectbox("Tahrir qilmoqchi bo'lgan qiymatni tanlang", list(values_dict.keys()))
        
        # Qiymatni yangilash
        if selected_value:
            value_id = values_dict[selected_value]
            new_value = st.number_input("Yangi qiymatni kiriting", value=1, min_value=0, max_value=1)
            if st.button("Yangilash", key="update_value", icon='🔄', use_container_width=True):
                if update_data("disease_symptoms", "value", new_value, "id", value_id):
                    st.success("✅ Qiymat yangilandi!")

            # Qiymatni o'chirish
            if st.button("O'chirish", key="delete_value", type='primary', icon='❌', use_container_width=True):
                if delete_data("disease_symptoms", "id", value_id):
                    st.success("✅ Qiymat o'chirildi!")
        
        # Qiymatni qo'shish
        new_value = st.number_input("Yangi qiymatni kiriting", value=1, min_value=0, max_value=1, key='new_value')
        if st.button("Qiymat qo'shish"):
            if new_value:
                with get_connection() as conn:
                    disease_id = st.selectbox("Kasallikni tanlang", [d[1] for d in diseases])
                    symptom_id = st.selectbox("Simptomni tanlang", [s[2] for s in symptoms])
                    save_symptom_value(conn, disease_id, symptom_id, new_value)
                st.success(f"✅ Qiymat qo'shildi!")


# Asosiy qism
st.markdown("# 💉Kasalliklar bilan ishlash oynasi")
tabs = st.tabs(["Ma'lumotlarni kiritish", "Ma'lumotlarni tahrirlash", "Excelga eksport qilish"])

with tabs[0]:  
    # 'st_tags' orqali kasallik nomlarini kiritish  
    kasallik_names = st_tags(  
        label='### Kasallik nomlarini kiriting...',  
        text="Har bir ma'lumot ENTER ni bosish orqali kiritiladi",  
        value=[], 
        suggestions=['Ревматоидный артрит', 'Подагрический артрит', 'Реактивный артрит'], 
        key="aljnddas"  
    )  
    
    if kasallik_names:  
        # Kasalliklar uchun simptomlar guruhi va simptomlarni kiritish  
        symptom_groups = []  
        symptoms = {}  

        for i, kasallik in enumerate(kasallik_names):  
            if kasallik:  
                with st.expander(f"{kasallik} uchun simptomlar guruhi va simptomlarni kiritish"):  
                    # Simptomlar guruhi kiritish uchun st_tags
                    group_names = st_tags(  
                        label=f'#### {kasallik} uchun simptomlar guruhi nomlarini kiriting',  
                        text="Har bir ma'lumot ENTER ni bosish orqali kiritiladi",  
                        value=[],  
                        key=f"group_tags_{i}"  
                    )
                    
                    for group_name in group_names:  
                        if group_name:  
                            symptoms[group_name] = st_tags(  
                                label=f"{group_name} uchun simptomlarni kiriting (vergul bilan ajrating)",  
                                text="Simptomlarni kiriting",  
                                value=[],  # Make sure this is defined correctly
                                key=f"symptoms_{i}_{group_name}"
                            )  

                    symptom_groups.append((kasallik, group_names))  

        # Simptomlarga mos ravishda 0 yoki 1 raqamlarini kiritish  
        symptom_matrix = []  

        for kasallik, groups in symptom_groups:  
            with st.expander(f"{kasallik} uchun simptomlar va checkboxlar"):  
                for group in groups:  
                    symptoms_text = symptoms.get(group, [])
                    symptom_list = symptoms_text  # `st_tags` natijasi ro'yxat

                    for symptom in symptom_list:  
                        # Checkbox qiymatini tekshirish  
                        value = st.checkbox(  
                            f"Kasallik: {kasallik} | Guruh: {group} | Simptom: {symptom} mavjudmi?",  
                            key=f"value_{kasallik}_{group}_{symptom}")  

                        # Agar checkbox tanlangan bo'lsa, qiymat 1, aks holda 0 bo'ladi  
                        symptom_matrix.append([kasallik, group, symptom, 1 if value else 0])  

        # Saqlash tugmasi  
        saqlash = st.button("Saqlash", disabled=not any(kasallik_names), icon='💡')  

        if saqlash:  
            try:  
                with get_connection() as conn:  
                    cursor = conn.cursor()  

                    # Simptomlar va kasalliklarga mos qiymatlarni bazaga saqlash  
                    for record in symptom_matrix:  
                        kasallik, group, symptom, value = record  

                        # Kasallikni olish yoki yaratish  
                        cursor.execute("SELECT id FROM diseases WHERE name=?", (kasallik,))  
                        disease_record = cursor.fetchone()  
                        if not disease_record:  
                            cursor.execute("INSERT INTO diseases (name) VALUES (?)", (kasallik,))  
                            conn.commit()  
                            disease_id = cursor.lastrowid  
                        else:  
                            disease_id = disease_record[0]  

                        # Simptomni olish yoki yaratish  
                        cursor.execute("SELECT id FROM symptom_groups WHERE disease_id=? AND group_name=?", (disease_id, group))  
                        group_record = cursor.fetchone()  
                        if not group_record:  
                            cursor.execute("INSERT INTO symptom_groups (disease_id, group_name) VALUES (?, ?)", (disease_id, group))  
                            conn.commit()  
                            group_id = cursor.lastrowid  
                        else:  
                            group_id = group_record[0]  

                        cursor.execute("SELECT id FROM symptoms WHERE group_id=? AND symptom_name=?", (group_id, symptom))  
                        symptom_record = cursor.fetchone()  
                        if not symptom_record:  
                            cursor.execute("INSERT INTO symptoms (group_id, symptom_name) VALUES (?, ?)", (group_id, symptom))  
                            conn.commit()  
                            symptom_id = cursor.lastrowid  
                        else:  
                            symptom_id = symptom_record[0]  

                        # Qiymatni saqlash (1 yoki 0)  
                        cursor.execute("""INSERT INTO disease_symptoms (disease_id, symptom_id, value)  
                                          VALUES (?, ?, ?)""", (disease_id, symptom_id, value))  
                        conn.commit()  

                st.success("✅ Kasalliklar va simptomlar muvaffaqiyatli saqlandi!")  
            except sqlite3.Error as e:  
                st.error(f"Xatolik yuz berdi: {str(e)}")

with tabs[1]:
    edit_tab()

def export_to_excel():
    diseases, groups, symptoms, values = get_all_data()
    
    # Excel yaratish
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Kasalliklar va Simptomlar"

    # Stil yaratish
    header_font = Font(name='Times New Roman', size=14, bold=True)
    cell_font = Font(name='Times New Roman', size=14)
    cell_alignment = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'))

    # Sarlavhalar
    headers = ['Simptomlar Guruhi', 'Simptom'] + [d[1] for d in diseases]
    for col, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.alignment = cell_alignment
        cell.border = border

    # Ma'lumotlarni to'ldirish
    current_row = 2
    for symptom in symptoms:
        worksheet.cell(row=current_row, column=1, value=symptom[3]).font = cell_font
        worksheet.cell(row=current_row, column=1).alignment = cell_alignment
        worksheet.cell(row=current_row, column=1).border = border
        
        worksheet.cell(row=current_row, column=2, value=symptom[2]).font = cell_font
        worksheet.cell(row=current_row, column=2).alignment = cell_alignment
        worksheet.cell(row=current_row, column=2).border = border
        
        for col, disease in enumerate(diseases, 3):
            cell = worksheet.cell(row=current_row, column=col)
            value = next((v[3] for v in values 
                        if v[1] == disease[0] and v[2] == symptom[0]), 0)
            cell.value = value
            cell.font = cell_font
            cell.alignment = cell_alignment
            cell.border = border
        
        current_row += 1
    # Ustunlarni kengaytirish
    for column in worksheet.columns:
        max_length = max(len(str(cell.value)) for cell in column)
        worksheet.column_dimensions[column[0].column_letter].width = max_length + 4
    return workbook

with tabs[2]:
    st.markdown("## Ma'lumotlar bazasidagi ma'lumotlar")
    # Ma'lumotlarni olish
    diseases, groups, symptoms, values = get_all_data()
    tabDiseases, tabGroups, tabSymptom, tabValues = st.tabs(['Kasalliklar', 'Simptom guruhlari', 'Simptomlar', 'Qiymatlar'])
    
    with tabDiseases:
        # Kasalliklar jadvalini ko'rsatish
        st.data_editor(pd.DataFrame(diseases, columns=["T/R", "Kasallik nomi"]))
    with tabGroups:
    # Simptom guruhlari jadvalini ko'rsatish
        st.data_editor(pd.DataFrame(groups, columns=["ID", "Kasallik ID", "Guruh nomi", "Kasallik nomi"]))

    with tabSymptom:
    # Simptomlar jadvalini ko'rsatish
        st.data_editor(pd.DataFrame(symptoms, columns=["ID", "Guruh ID", "Simptom nomi", "Guruh nomi"]))

    with tabValues:
    # Qiymatlar jadvalini ko'rsatish
        st.data_editor(pd.DataFrame(values, columns=["ID", "Kasallik ID", "Simptom ID", "Qiymat", "Kasallik nomi", "Simptom nomi", "Guruh nomi"]))
    
    # Excelga eksport qilish funksiyasini chaqirish
    workbook = export_to_excel()
    excel_file = BytesIO()
    workbook.save(excel_file)
    excel_file.seek(0)
    st.download_button(
        label="Jadval ko'rinishida yuklash",
        data=excel_file,
        file_name="kasalliklar_va_simptomlar.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        icon='💾',
        type='primary'
    )
