import sqlite3
import streamlit as st
import pandas as pd
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from contextlib import contextmanager

# Thread-safe database connection management
@contextmanager
def get_connection():
    conn = None
    try:
        conn = sqlite3.connect('data/diagnosis_data.db', timeout=10, check_same_thread=False)
        yield conn
    finally:
        if conn:
            conn.close()

# Ma'lumotlarni olish funksiyasi
def get_all_data():
    with get_connection() as conn:
        cursor = conn.cursor()
        
        diseases = cursor.execute("SELECT id, name FROM diseases").fetchall()
        groups = cursor.execute("""
            SELECT sg.id, sg.disease_id, sg.group_name, d.name as disease_name
            FROM symptom_groups sg
            JOIN diseases d ON sg.disease_id = d.id
        """).fetchall()
        symptoms = cursor.execute("""
            SELECT s.id, s.group_id, s.symptom_name, sg.group_name
            FROM symptoms s
            JOIN symptom_groups sg ON s.group_id = sg.id
        """).fetchall()
        values = cursor.execute("""
            SELECT ds.id, ds.disease_id, ds.symptom_id, ds.value,
                   d.name as disease_name, s.symptom_name, sg.group_name
            FROM disease_symptoms ds
            JOIN diseases d ON ds.disease_id = d.id
            JOIN symptoms s ON ds.symptom_id = s.id
            JOIN symptom_groups sg ON s.group_id = sg.id
        """).fetchall()
        
        return diseases, groups, symptoms, values

def update_data(table, field, value, condition_field, condition_value):
    with get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute(f"UPDATE {table} SET {field}=? WHERE {condition_field}=?", 
                      (value, condition_value))
        conn.commit()

def delete_data(table, condition_field, condition_value):
    with get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute(f"DELETE FROM {table} WHERE {condition_field}=?", (condition_value,))
        conn.commit()

# Session state orqali ma'lumotlarni saqlash
if 'cached_data' not in st.session_state:
    st.session_state.cached_data = get_all_data()

def refresh_data():
    st.session_state.cached_data = get_all_data()

# Tahrirlash oynasi
def edit_tab():
    st.markdown("## :rainbow[Ma'lumotlarni tahrirlash]")
    
    edit_type = st.selectbox("Nimani tahrirlash kerak?", 
                            ["Kasalliklar", "Simptom guruhlari", "Simptomlar", "Qiymatlar"])
    
    diseases, groups, symptoms, values = st.session_state.cached_data
    
    if edit_type == "Kasalliklar":
        disease_dict = {d[1]: d[0] for d in diseases}
        selected_disease = st.selectbox("Kasallikni tanlang", disease_dict.keys())
        
        if selected_disease:
            new_name = st.text_input("Yangi nom", value=selected_disease)
            col1, col2 = st.columns(2)
            
            if col1.button("Yangilash"):
                update_data("diseases", "name", new_name, "id", disease_dict[selected_disease])
                refresh_data()
                st.success("Kasallik nomi yangilandi!")
            
            if col2.button("O'chirish"):
                delete_data("diseases", "id", disease_dict[selected_disease])
                refresh_data()
                st.success("Kasallik o'chirildi!")

    elif edit_type == "Simptom guruhlari":
        group_dict = {f"{g[2]} ({g[3]})": g[0] for g in groups}
        selected_group = st.selectbox("Simptom guruhini tanlang", group_dict.keys())
        
        if selected_group:
            new_name = st.text_input("Yangi nom", value=selected_group.split(" (")[0])
            col1, col2 = st.columns(2)
            
            if col1.button("Yangilash"):
                update_data("symptom_groups", "group_name", new_name, 
                          "id", group_dict[selected_group])
                refresh_data()
                st.success("Guruh nomi yangilandi!")
            
            if col2.button("O'chirish"):
                delete_data("symptom_groups", "id", group_dict[selected_group])
                refresh_data()
                st.success("Guruh o'chirildi!")

    # Simptomlar va Qiymatlar uchun kodlar xuddi shunday davom etadi...
    # (Avvalgi koddan nusxa oling, faqat refresh_data() qo'shing)

# Export funksiyasi (o'zgarishsiz qoladi)
def export_to_excel():
    diseases, groups, symptoms, values = get_all_data()
    
    # Excel yaratish
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Kasalliklar va Simptomlar"

    # Stil yaratish
    header_font = Font(name='Times New Roman', size=14, bold=True)
    cell_font = Font(name='Times New Roman', size=14)
    cell_alignment = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                   top=Side(style='thin'), bottom=Side(style='thin'))

    # Sarlavhalar
    headers = ['Simptomlar Guruhi', 'Simptom'] + [d[1] for d in diseases]
    for col, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.alignment = cell_alignment
        cell.border = border

    # Ma'lumotlarni to'ldirish
    current_row = 2
    for symptom in symptoms:
        worksheet.cell(row=current_row, column=1, value=symptom[3]).font = cell_font
        worksheet.cell(row=current_row, column=1).alignment = cell_alignment
        worksheet.cell(row=current_row, column=1).border = border
        
        worksheet.cell(row=current_row, column=2, value=symptom[2]).font = cell_font
        worksheet.cell(row=current_row, column=2).alignment = cell_alignment
        worksheet.cell(row=current_row, column=2).border = border
        
        for col, disease in enumerate(diseases, 3):
            cell = worksheet.cell(row=current_row, column=col)
            value = next((v[3] for v in values 
                        if v[1] == disease[0] and v[2] == symptom[0]), 0)
            cell.value = value
            cell.font = cell_font
            cell.alignment = cell_alignment
            cell.border = border
        
        current_row += 1

    # Ustunlarni kengaytirish
    for column in worksheet.columns:
        max_length = max(len(str(cell.value)) for cell in column)
        worksheet.column_dimensions[column[0].column_letter].width = max_length + 4

    return workbook


# Asosiy qism
st.markdown("# 💉:rainbow[Kasalliklar bilan ishlash oynasi]")
tabs = st.tabs(["Ma'lumotlarni kiritish", "Ma'lumotlarni tahrirlash", "Excelga eksport qilish"])

with tabs[0]:
    st.markdown("## :rainbow[Kasalliklar va Simptomlarni kiritish]")
    
    kasalliklar_soni = st.number_input("Nechta kasallik nomini kiritasiz?", 1, 10, 1)
    kasalliklar_nomlari = [st.text_input(f"{i+1}-Kasallik nomini kiriting", key=f'kasallik_{i}') for i in range(kasalliklar_soni)]

    # Kasalliklar uchun simptomlar guruhi va simptomlarni kiritish
    symptom_groups = []
    symptoms = {}

    for i, kasallik in enumerate(kasalliklar_nomlari):
        if kasallik:
            st.markdown(f"### {kasallik} uchun simptomlar guruhi kiritish")
            num_groups = st.number_input(f"{kasallik} uchun simptomlar guruhi sonini kiriting", 0, 5, 1, key=f"num_groups_{i}")
            groups = []
            for j in range(num_groups):
                group_name = st.text_input(f"{j+1}-Simptomlar guruhi nomini kiriting", key=f"group_{i}_{j}")
                if group_name:
                    groups.append(group_name)
                    symptoms[group_name] = st.text_area(f"{group_name} uchun simptomlarni kiriting (vergul bilan ajrating)", key=f"symptoms_{i}_{j}")
            symptom_groups.append((kasallik, groups))

    # Simptomlarga mos ravishda 0 yoki 1 raqamlarini kiritish
    symptom_matrix = []

    for kasallik, groups in symptom_groups:
        for group in groups:
            symptoms_text = symptoms[group]
            symptom_list = [symptom.strip() for symptom in symptoms_text.split(',')]
            
            # Har bir simptom uchun checkbox (0 yoki 1)
            for symptom in symptom_list:
                if symptom:
                    value = st.checkbox(f"Kasallik: {kasallik} | Guruh: {group} | Simptom: {symptom} mavjudmi?", key=f"value_{kasallik}_{group}_{symptom}")
                    symptom_matrix.append([kasallik, group, symptom, 1 if value else 0])

    # Saqlash tugmasi
    saqlash = st.button("Saqlash", disabled=not kasalliklar_nomlari)

    if saqlash:
        st.write(f"Saqlash muvaffaqiyatli bajarildi. Jami {len(kasalliklar_nomlari)} ta kasallik kiritildi.")
        
        # Kasalliklarni va simptomlarni saqlash
        conn = get_connection()
        cursor = conn.cursor()
        
        for kasallik in kasalliklar_nomlari:
            if kasallik:
                # Kasallik nomi bazada mavjudligini tekshirish
                cursor.execute("SELECT id FROM diseases WHERE name=?", (kasallik,))
                disease_id = cursor.fetchone()
                
                if not disease_id:
                    cursor.execute("INSERT INTO diseases (name) VALUES (?)", (kasallik,))
                    disease_id = cursor.lastrowid
                else:
                    disease_id = disease_id[0]

                # Har bir kasallik uchun simptomlar guruhi saqlash
                for group in symptoms:
                    # Simptomlar guruhi bazada mavjudligini tekshirish
                    cursor.execute("SELECT id FROM symptom_groups WHERE disease_id=? AND group_name=?", (disease_id, group))
                    group_id = cursor.fetchone()
                    
                    if not group_id:
                        cursor.execute("INSERT INTO symptom_groups (disease_id, group_name) VALUES (?, ?)", (disease_id, group))
                        group_id = cursor.lastrowid
                    else:
                        group_id = group_id[0]

                        # Har bir simptomni saqlash
                        for symptom in symptoms[group].split(','):
                            symptom = symptom.strip()
                            if symptom:
                                cursor.execute("INSERT INTO symptoms (group_id, symptom_name) VALUES (?, ?)", (group_id, symptom))

                # Simptomlar jadvaliga 0 yoki 1 qiymatlarni saqlash
                data_to_insert = []
                for kasallik, group, symptom, value in symptom_matrix:
                    disease_id = cursor.execute("SELECT id FROM diseases WHERE name=?", (kasallik,)).fetchone()
                    if disease_id:
                        disease_id = disease_id[0]
                        group_id = cursor.execute("SELECT id FROM symptom_groups WHERE disease_id=? AND group_name=?", (disease_id, group)).fetchone()
                        
                        if group_id:
                            group_id = group_id[0]
                            symptom_id = cursor.execute("SELECT id FROM symptoms WHERE group_id=? AND symptom_name=?", (group_id, symptom)).fetchone()
                            
                            if symptom_id:
                                symptom_id = symptom_id[0]
                                
                                # Ma'lumotlar bazasida mavjudligini tekshirish va saqlash
                                cursor.execute("""
                                    INSERT OR IGNORE INTO disease_symptoms (disease_id, symptom_id, value) 
                                    VALUES (?, ?, ?)
                                """, (disease_id, symptom_id, value))

        conn.commit()  # O'zgarishlarni saqlash
        conn.close()  # Ulanishni yopish
        st.success("Ma'lumotlar muvaffaqiyatli saqlandi!")

with tabs[1]:
    edit_tab()

with tabs[2]:
    st.markdown("## :rainbow[Excelga eksport qilish]")
    col1, col2 = st.columns(2)
    
    if col1.button("Excelga eksport qilish", icon='⏳', type='primary'):
        workbook = export_to_excel()
        
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        
        col2.download_button(
            label="Excel faylini yuklab olish",
            data=output.getvalue(),
            file_name="kasalliklar_va_simptomlar.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            icon='💾'
        )